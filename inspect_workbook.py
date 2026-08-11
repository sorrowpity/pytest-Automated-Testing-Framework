from openpyxl import load_workbook

wb = load_workbook(r'F:\software_learn\python\mihayo_learn\data\data1.xlsx')
ws = wb['Sheet1']
print('max_row', ws.max_row, 'max_column', ws.max_column)
print('headers', [c.value for c in ws[2]])
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    print(row)
